"""
사물함 배정 모듈

좌석 배정 결과(seat_result.csv)를 읽어, 각 열람실에 해당하는 사물함 번호를 배정합니다.

배정 흐름:
  1. config.yaml에서 열람실 → 사물함 매핑 정보를 로드
  2. 좌석 배정 결과를 읽어 학생 목록을 만듦
  3. 학생 순서를 랜덤 셔플
  4. 각 학생의 열람실에 맞는 사물함을 순차 배정 (overflow 자동 처리)
  5. 결과를 CSV 파일로 저장
"""

import random
import csv
import argparse
from collections import defaultdict
from openpyxl import Workbook, load_workbook

from config import load_config
# [2025.8.] 추가 배정 시 생방송 진행하는 대신 input file 기반 시드 고정
from seat import get_seed_from_file


# ============================================================
# 사물함 상태 관리
# ============================================================

def build_locker_state(config):
    """
    config의 locker_mapping을 기반으로 사물함 배정 상태를 초기화합니다.

    각 열람실의 사물함은 리스트로 정의되어 있어, 첫 번째 사물함이 가득 차면
    다음 사물함으로 자동 overflow됩니다.

    Returns:
        locker_state: 각 사물함의 현재 배정 상태
            { (열람실, 순번): { 'location': 위치명, 'start': 시작번호, 'current': 다음배정번호, 'end': 끝번호 } }
        room_to_lockers: 열람실 → 사물함 키 목록 (순서대로 채움)
            { 열람실: [(열람실, 0), (열람실, 1), ...] }
    """
    locker_state = {}
    room_to_lockers = {}

    for room, info in config['locker_mapping'].items():
        room_to_lockers[room] = []
        for idx, locker in enumerate(info['lockers']):
            key = (room, idx)
            locker_state[key] = {
                'location': locker['location'],
                'start': locker['start'],
                'current': locker['start'],
                'end': locker['end'],
            }
            room_to_lockers[room].append(key)

    return locker_state, room_to_lockers


def assign_locker(room, locker_state, room_to_lockers):
    """
    주어진 열람실에 대해 사물함 번호를 하나 배정합니다.

    lockers 리스트를 순서대로 순회하며, 첫 번째 사물함이 가득 차면 다음으로 overflow합니다.
    예) 15동 404호(칸막이) → 404(A) 55~66번 다 차면 → 404(B) 1~150번으로 이동

    Returns: [사물함위치, 번호] 또는 None (모두 가득 찬 경우)
    """
    if room not in room_to_lockers:
        return None

    for key in room_to_lockers[room]:
        state = locker_state[key]
        if state['current'] <= state['end']:
            result = [state['location'], state['current']]
            state['current'] += 1
            return result

    return None


def validate_locker_capacity(locker_state):
    """사물함 번호가 설정된 용량을 초과했는지 검증합니다."""
    for (room, idx), state in locker_state.items():
        if state['current'] > state['end'] + 1:
            print(f"[-] {state['location']} 사물함 넘버 초과 (열람실: {room})")


def load_indices_from_existing(file_path, config):
    """
    기존 seat_locker_result.csv를 읽어, 이미 배정된 사물함 번호만큼
    사물함 상태의 current 인덱스를 전진시킵니다 (추가 배정 시 사용).
    """
    locker_state, room_to_lockers = build_locker_state(config)

    try:
        with open(file_path, mode='rt', encoding='UTF-8') as csvfile:
            csvreader = csv.reader(csvfile)
            next(csvreader)  # header skip
            for row in csvreader:
                locker_location = row[4]
                locker_num = int(row[5])

                # 매칭되는 사물함 상태를 찾아 인덱스 전진
                for key, state in locker_state.items():
                    if (state['location'] == locker_location and
                            state['start'] <= locker_num <= state['end']):
                        state['current'] = max(state['current'], locker_num + 1)
                        break
    except FileNotFoundError:
        print(f"[!] 기존 파일 {file_path} 없음. 기본 인덱스로 진행하나, 반드시 수작업 필요")

    return locker_state, room_to_lockers


# ============================================================
# 메인 실행
# ============================================================

def main(mode="normal"):
    """좌석 배정 결과를 읽어 사물함을 배정합니다."""
    config = load_config()
    paths = config['paths']

    # 좌석배치 완료된 파일에서 학생 정보 로드
    if mode == "normal":
        file_path = paths['output_result']
        locker_state, room_to_lockers = build_locker_state(config)
    else:
        file_path = paths['output_result_additional']
        # 추가 배정: 입력 파일 해시 기반 시드 → 동일 입력이면 동일 결과 보장
        random.seed(get_seed_from_file(file_path))
        locker_state, room_to_lockers = load_indices_from_existing(
            paths['output_locker_result'], config)

    students = []
    with open(file_path, mode='rt', encoding='UTF-8', newline='') as csvfile:
        csvreader = csv.reader(csvfile)
        next(csvreader)  # 헤더 skip
        for row in csvreader:
            students.append(row)
    random.shuffle(students)

    # 각 학생에게 사물함 배정
    # student: [이름, 학번뒤2자리, 열람실, 좌석번호, 1지망배정여부]
    # locker:  [사물함위치, 사물함번호]
    # 출력 순서: 이름, 학번뒤2자리, 열람실, 좌석번호, 사물함, 사물함번호, 1지망배정여부
    result = []
    failed = defaultdict(int)

    for student in students:
        room = student[2]  # 열람실명
        locker = assign_locker(room, locker_state, room_to_lockers)
        if locker is None:
            failed[room] += 1
            continue
        first_pref = student[4]  # 1지망배정여부 (O/X)
        result.append(student[:4] + locker + [first_pref])

    # 검증
    validate_locker_capacity(locker_state)

    if len(result) != len(students):
        print("[-] 전체 숫자 안맞음. 데이터 오타 확인할 것")
        print(f"[-] 배정 성공: {len(result)}, 전체: {len(students)}, 실패: {sum(failed.values())}")

    if failed:
        for room, count in failed.items():
            print(f"[-] 열람실: {room}, 실패 횟수: {count}")

    # 결과 CSV 저장
    HEADER = ["이름", "학번뒤2자리", "열람실", "좌석번호", "사물함", "사물함번호", "1지망배정여부"]

    def write_locker_csv(filepath, rows, write_header=True):
        mode = 'wt' if write_header else 'at'
        with open(filepath, mode=mode, encoding='utf-8') as file:
            if write_header:
                file.write(",".join(HEADER) + "\n")
            for r in rows:
                file.write(f"{r[0]},{r[1]},{r[2]},{r[3]},{r[4]},{r[5]},{r[6]}\n")
        print(f'[+] 좌석 및 사물함 배치 결과 저장 경로: {filepath}')

    def _save_xlsx(wb, filepath):
        """xlsx 저장. 실패해도 전체 프로세스를 중단하지 않는다."""
        try:
            wb.save(filepath)
            print(f'[+] 좌석 및 사물함 배치 결과 저장 경로: {filepath}')
        except PermissionError:
            print(f'[!] {filepath} 저장 실패: 파일이 다른 프로그램(Excel 등)에서 열려 있습니다.')
            print(f'    CSV 파일은 정상 저장되었으니, xlsx는 파일을 닫고 다시 실행해주세요.')
        except Exception as e:
            print(f'[!] {filepath} 저장 실패: {e}')
            print(f'    CSV 파일은 정상 저장되었으니, xlsx는 다시 실행해주세요.')

    def write_locker_xlsx(filepath, rows):
        """결과를 xlsx 파일로 저장합니다 (새로 생성)."""
        wb = Workbook()
        ws = wb.active
        ws.append(HEADER)
        for r in rows:
            ws.append(list(r))
        _save_xlsx(wb, filepath)

    def append_locker_xlsx(filepath, rows):
        """기존 xlsx 파일에 행을 추가합니다."""
        try:
            wb = load_workbook(filepath)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(HEADER)
        for r in rows:
            ws.append(list(r))
        _save_xlsx(wb, filepath)

    if mode == "normal":
        write_locker_csv(paths['output_locker_result'], result)
        write_locker_xlsx(paths['output_locker_result_xlsx'], result)
    else:
        # 추가 배정: 기존 파일에 append + 별도 추가분 파일 생성
        write_locker_csv(paths['output_locker_result'], result, write_header=False)
        append_locker_xlsx(paths['output_locker_result_xlsx'], result)
        write_locker_csv(paths['output_locker_result_additional'], result)
        write_locker_xlsx(paths['output_locker_result_additional_xlsx'], result)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=["normal", "add"], default="normal")
    args = parser.parse_args()
    main(mode=args.mode)
